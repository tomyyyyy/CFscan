import openpyxl
import json


class Fingerprint(object):
    def __init__(self,script_path):
        self.script_path = script_path

        # 读取指纹数据库表单 
        wb = openpyxl.load_workbook(F"{self.script_path}/resourses/摄像头指纹库.xlsx")
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]

        # 建立指纹数据字典
        eq = {'camera':{}}
        i = 1
        while(True):
            i += 1
            # 获取品牌
            brand = str(ws['A' + str(i)].value)

            if brand != 'None':
                # 获取品牌对应的型号列表
                modellist = ws['B' + str(i)].value.replace('\'', '').split(', ')
                # 填充字典
                eq['camera'][brand] = modellist        
            else:
                    break

        # 指纹识别
        with open(F"{self.script_path}/json/fo.json", 'r',  encoding = 'utf-8') as f:
            for line in f.readlines():
                line = line.strip()   # 使用strip函数去除空行
                if len(line) != 0:

                    thing = {'ip':'','equipment':'','brand':'', 'model':''}
                    # 先提取ip
                    lines = json.loads(line)
                    ip = lines['ip']
                    print(ip)
                    data = lines['data']
                    # 查该ip设备类型
                    for e in eq.keys():   
                        #if any(e in st for st in data):
                        thing['ip'] = ip
                        thing['equipment'] = e
                        # 查品牌
                        for b in eq[e].keys():
                                if b == 'Brickcom' or b == 'Huawei':
                                        if any( b.lower() in st.lower() for st in data):
                                                thing['brand'] = b
                                                # 查型号
                                                for m in eq[e][b]:
                                                        if any(m.lower() in st.lower() for st in data):
                                                                thing['model'] = m
                                else: 
                                        for m in eq[e][b]:
                                                if any(m.lower() in st.lower() for st in data) and len(m) > 3:
                                                        thing['brand'] = b
                                                        thing['model'] = m
                    if thing['brand'] != '':
                        with open(F"{self.script_path}/json/end.json","a", encoding = 'utf-8') as f:
                            f.write(json.dumps(thing) + '\n')





        



